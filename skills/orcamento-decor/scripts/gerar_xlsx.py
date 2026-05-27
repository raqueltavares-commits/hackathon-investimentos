"""Monta um .xlsx multi-aba (uma aba por tipologia + Resumo) a partir do JSON de montar_orcamento.

Uso:
  python montar_orcamento.py ... > memoriais.json
  python gerar_xlsx.py --memoriais memoriais.json --spot "Bonito Spot" --saida saida.xlsx
"""
import argparse
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

NAVY = "FF000C3C"
bold = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFFFF")
navy_fill = PatternFill("solid", fgColor=NAVY)


def _aba(ws, rows):
    for r in rows:
        ws.append(r)
    for c in ws[2]:
        c.font = bold_white
        c.fill = navy_fill
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22
    for col in ("H", "I", "J", "K", "L"):
        ws.column_dimensions[col].width = 16


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    p = argparse.ArgumentParser()
    p.add_argument("--memoriais", type=Path, required=True)
    p.add_argument("--spot", required=True)
    p.add_argument("--saida", type=Path, required=True)
    args = p.parse_args()

    data = json.loads(args.memoriais.read_text(encoding="utf-8"))
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumo"
    ws.append([f"{args.spot} — Orçamento Decor (Plus)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["TIPOLOGIA", "DESCRIÇÃO", "CUSTO ESTIMADO (sem jacuzzi)", "COM JACUZZI"])
    for c in ws[3]:
        c.font = bold_white
        c.fill = navy_fill
    for m in data["memoriais"]:
        ws.append([m["tipologia"], m["descricao"], m["total_geral"], m["total_com_jacuzzi"]])
        ws.cell(row=ws.max_row, column=3).number_format = 'R$ #,##0.00'
        ws.cell(row=ws.max_row, column=4).number_format = 'R$ #,##0.00'
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 18

    for m in data["memoriais"]:
        aba = wb.create_sheet(title=f"Tipologia {m['tipologia']}")
        _aba(aba, m["rows"])

    wb.save(args.saida)
    print(f"xlsx salvo: {args.saida} | abas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
